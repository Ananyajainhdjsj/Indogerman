import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_supply_chain_excel(filename='supply_chain_data.xlsx'):
    """
    Create user-friendly Excel file with expandable rows for multiple facilities.
    """
    
    print("Creating Excel file with multiple sheets...")
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    instruction_font = Font(italic=True, size=10, color="0000FF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================================================
    # SHEET 1: PLANTS (Production Units)
    # ============================================================================
    print("  Creating sheet 1: Plants")
    ws_plants = wb.create_sheet("Plants")
    
    plants_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    ws_plants.append(['PRODUCTION PLANTS / MANUFACTURING UNITS'])
    ws_plants.append(['Add as many rows as needed for your production facilities'])
    ws_plants.append([''])
    ws_plants.append(['Plant Code', 'Plant Name', 'Production Cost (€/KWp)', 'Capacity (KWp)', 'Emissions Factor (kg CO2e/KWp)'])
    ws_plants.append(['P1', 'Berlin Production Plant', 140.0, 10000.0, 450.0])
    ws_plants.append(['P2', 'Munich Production Plant', 145.0, 8000.0, 460.0])
    ws_plants.append(['', '', '', '', ''])  # Empty row for adding more
    
    # Format
    ws_plants['A1'].font = Font(bold=True, size=14)
    ws_plants['A2'].font = instruction_font
    for cell in ws_plants[4]:
        cell.font = header_font
        cell.fill = plants_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 2: CUSTOMERS (Market Zones)
    # ============================================================================
    print("  Creating sheet 2: Customers")
    ws_customers = wb.create_sheet("Customers")
    
    customers_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    ws_customers.append(['CUSTOMERS / MARKET ZONES'])
    ws_customers.append(['Add as many customer zones as needed'])
    ws_customers.append([''])
    ws_customers.append(['Customer Code', 'Customer Name', 'Product Type', 'Demand (KWp)', 'Returns (KWp)'])
    ws_customers.append(['C1', 'Hamburg Market', 'Monocrystalline', 1000.0, 800.0])
    ws_customers.append(['C2', 'Frankfurt Market', 'Monocrystalline', 1000.0, 800.0])
    ws_customers.append(['C3', 'Stuttgart Market', 'Monocrystalline', 1200.0, 900.0])
    ws_customers.append(['', '', '', '', ''])  # Empty row
    
    # Format
    ws_customers['A1'].font = Font(bold=True, size=14)
    ws_customers['A2'].font = instruction_font
    for cell in ws_customers[4]:
        cell.font = header_font
        cell.fill = customers_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 3: COLLECTION CENTERS
    # ============================================================================
    print("  Creating sheet 3: Collection_Centers")
    ws_collection = wb.create_sheet("Collection_Centers")
    
    collection_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    ws_collection.append(['COLLECTION CENTERS'])
    ws_collection.append(['Add collection facilities - leave Fixed Cost = 0 if already open'])
    ws_collection.append([''])
    ws_collection.append(['Code', 'Name', 'Collection Cost (€/KWp)', 'Fixed Cost (€)', 'Capacity (kg)', 'Emissions Factor (kg CO2e/KWp)'])
    ws_collection.append(['O1', 'Hamburg Collection Center', 8.0, 15000.0, 100000.0, 5.0])
    ws_collection.append(['O2', 'Frankfurt Collection Center', 8.5, 15000.0, 120000.0, 5.0])
    ws_collection.append(['', '', '', '', '', ''])
    
    # Format
    ws_collection['A1'].font = Font(bold=True, size=14)
    ws_collection['A2'].font = instruction_font
    for cell in ws_collection[4]:
        cell.font = header_font
        cell.fill = collection_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 4: REFURBISHMENT CENTERS
    # ============================================================================
    print("  Creating sheet 4: Refurbishment_Centers")
    ws_refurb = wb.create_sheet("Refurbishment_Centers")
    
    refurb_header_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    
    ws_refurb.append(['REFURBISHMENT CENTERS'])
    ws_refurb.append(['Add refurbishment facilities'])
    ws_refurb.append([''])
    ws_refurb.append(['Code', 'Name', 'Refurbishing Cost (€/KWp)', 'Fixed Cost (€)', 'Capacity (kg)', 'Yield (%)', 'Emissions Factor (kg CO2e/KWp)'])
    ws_refurb.append(['F1', 'Berlin Refurbishment', 25.0, 25000.0, 50000.0, 0.90, 30.0])
    ws_refurb.append(['F2', 'Munich Refurbishment', 26.0, 25000.0, 45000.0, 0.88, 32.0])
    ws_refurb.append(['', '', '', '', '', '', ''])
    
    # Format
    ws_refurb['A1'].font = Font(bold=True, size=14)
    ws_refurb['A2'].font = instruction_font
    for cell in ws_refurb[4]:
        cell.font = header_font
        cell.fill = refurb_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 5: RECYCLING CENTERS
    # ============================================================================
    print("  Creating sheet 5: Recycling_Centers")
    ws_recycle = wb.create_sheet("Recycling_Centers")
    
    recycle_header_fill = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
    
    ws_recycle.append(['RECYCLING CENTERS'])
    ws_recycle.append(['Add recycling facilities'])
    ws_recycle.append([''])
    ws_recycle.append(['Code', 'Name', 'Recycling Cost (€/kg)', 'Fixed Cost (€)', 'Capacity (kg)', 'Efficiency (%)', 'Emissions Factor (kg CO2e/kg)'])
    ws_recycle.append(['R1', 'Hamburg Recycling', 0.60, 30000.0, 50000.0, 0.95, 1.5])
    ws_recycle.append(['R2', 'Frankfurt Recycling', 0.62, 30000.0, 55000.0, 0.94, 1.5])
    ws_recycle.append(['', '', '', '', '', '', ''])
    
    # Format
    ws_recycle['A1'].font = Font(bold=True, size=14)
    ws_recycle['A2'].font = instruction_font
    for cell in ws_recycle[4]:
        cell.font = header_font
        cell.fill = recycle_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 6: LANDFILLS
    # ============================================================================
    print("  Creating sheet 6: Landfills")
    ws_landfill = wb.create_sheet("Landfills")
    
    landfill_header_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    
    ws_landfill.append(['LANDFILLS / DISPOSAL SITES'])
    ws_landfill.append(['Add landfill locations'])
    ws_landfill.append([''])
    ws_landfill.append(['Code', 'Name', 'Disposal Cost (€/kg)', 'Emissions Factor (kg CO2e/kg)'])
    ws_landfill.append(['L1', 'North Landfill', 0.15, 0.5])
    ws_landfill.append(['L2', 'South Landfill', 0.16, 0.5])
    ws_landfill.append(['', '', '', ''])
    
    # Format
    ws_landfill['A1'].font = Font(bold=True, size=14)
    ws_landfill['A2'].font = instruction_font
    for cell in ws_landfill[4]:
        cell.font = header_font
        cell.fill = landfill_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 7: SECONDARY MARKETS
    # ============================================================================
    print("  Creating sheet 7: Secondary_Markets")
    ws_secondary = wb.create_sheet("Secondary_Markets")
    
    secondary_header_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
    
    ws_secondary.append(['SECONDARY MARKETS / MATERIAL BUYERS'])
    ws_secondary.append(['Add buyers for recycled materials'])
    ws_secondary.append([''])
    ws_secondary.append(['Code', 'Name', 'Location'])
    ws_secondary.append(['S1', 'Material Buyer North', 'Hamburg'])
    ws_secondary.append(['S2', 'Material Buyer South', 'Munich'])
    ws_secondary.append(['', '', ''])
    
    # Format
    ws_secondary['A1'].font = Font(bold=True, size=14)
    ws_secondary['A2'].font = instruction_font
    for cell in ws_secondary[4]:
        cell.font = header_font
        cell.fill = secondary_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 8: DISTANCE MATRIX
    # ============================================================================
    print("  Creating sheet 8: Distance_Matrix")
    ws_dist = wb.create_sheet("Distance_Matrix")
    
    dist_header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    
    ws_dist.append(['DISTANCE MATRIX'])
    ws_dist.append(['Add all routes between facilities (in km)'])
    ws_dist.append([''])
    ws_dist.append(['From', 'To', 'Distance (km)'])
    ws_dist.append(['P1', 'C1', 50.0])
    ws_dist.append(['P1', 'C2', 75.0])
    ws_dist.append(['P2', 'C1', 80.0])
    ws_dist.append(['P2', 'C2', 60.0])
    ws_dist.append(['C1', 'O1', 30.0])
    ws_dist.append(['C2', 'O1', 40.0])
    ws_dist.append(['O1', 'F1', 45.0])
    ws_dist.append(['O1', 'R1', 35.0])
    ws_dist.append(['O1', 'L1', 55.0])
    ws_dist.append(['F1', 'P1', 50.0])
    ws_dist.append(['R1', 'S1', 25.0])
    ws_dist.append(['', '', ''])
    
    # Format
    ws_dist['A1'].font = Font(bold=True, size=14)
    ws_dist['A2'].font = instruction_font
    for cell in ws_dist[4]:
        cell.font = header_font
        cell.fill = dist_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 9: REVENUES
    # ============================================================================
    print("  Creating sheet 9: Revenues")
    ws_rev = wb.create_sheet("Revenues")
    
    rev_header_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    
    ws_rev.append(['REVENUE DATA'])
    ws_rev.append([''])
    ws_rev.append(['A. PRODUCT REVENUES'])
    ws_rev.append(['Product Type', 'Revenue Type', 'Revenue (€/KWp)'])
    ws_rev.append(['Monocrystalline', 'Reuse', 90.0])
    ws_rev.append(['Monocrystalline', 'Refurbished', 110.0])
    ws_rev.append([''])
    ws_rev.append(['B. MATERIAL REVENUES (from Recycling)'])
    ws_rev.append(['Material', 'Revenue (€/kg)'])
    ws_rev.append(['Glass', 0.08])
    ws_rev.append(['Aluminum', 1.80])
    ws_rev.append(['Silicon', 12.0])
    ws_rev.append(['Plastic', 0.15])
    ws_rev.append(['Copper', 6.50])
    
    # Format
    ws_rev['A1'].font = Font(bold=True, size=14)
    ws_rev['A3'].font = Font(bold=True, size=11)
    ws_rev['A8'].font = Font(bold=True, size=11)
    for cell in ws_rev[4]:
        cell.font = header_font
        cell.fill = rev_header_fill
        cell.border = thin_border
    for cell in ws_rev[9]:
        cell.font = header_font
        cell.fill = rev_header_fill
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 10: MATERIALS
    # ============================================================================
    print("  Creating sheet 10: Materials")
    ws_materials = wb.create_sheet("Materials")
    
    materials_header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    ws_materials.append(['MATERIAL COMPOSITION'])
    ws_materials.append(['Material content per KWp of solar panel'])
    ws_materials.append([''])
    ws_materials.append(['Material', 'Quantity (kg/KWp)'])
    ws_materials.append(['Glass', 8.0])
    ws_materials.append(['Aluminum', 1.5])
    ws_materials.append(['Silicon', 0.5])
    ws_materials.append(['Plastic', 0.8])
    ws_materials.append(['Copper', 0.2])
    
    # Format
    ws_materials['A1'].font = Font(bold=True, size=14)
    ws_materials['A2'].font = instruction_font
    for cell in ws_materials[4]:
        cell.font = header_font
        cell.fill = materials_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 11: PARAMETERS
    # ============================================================================
    print("  Creating sheet 11: Parameters")
    ws_params = wb.create_sheet("Parameters")
    
    params_header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    ws_params.append(['MODEL PARAMETERS'])
    ws_params.append([''])
    ws_params.append(['Parameter', 'Value', 'Unit', 'Description'])
    ws_params.append(['Transport Cost', 0.004, '€/kg-km', 'Cost per kg per km'])
    ws_params.append(['Penalty Cost', 20000.0, '€/KWp', 'Penalty for unmet demand'])
    ws_params.append(['Transport Emissions', 0.00006, 'kg CO2e/kg-km', 'Emissions per kg per km'])
    ws_params.append(['Panel Weight', 11.0, 'kg/KWp', 'Weight of solar panel'])
    ws_params.append(['Reuse Capacity', 0.20, 'ratio', 'Max % of returns that can be reused'])
    ws_params.append(['Refurbishment Capacity', 0.40, 'ratio', 'Max % of returns that can be refurbished'])
    ws_params.append(['Epsilon Limit', 50000.0, 'kg CO2e', 'Maximum allowed emissions'])
    ws_params.append(['Minimize Emissions Only', 'FALSE', 'TRUE/FALSE', 'TRUE to minimize emissions, FALSE to minimize cost'])
    
    # Format
    ws_params['A1'].font = Font(bold=True, size=14)
    for cell in ws_params[3]:
        cell.font = header_font
        cell.fill = params_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # SHEET 12: PRODUCT TYPES
    # ============================================================================
    print("  Creating sheet 12: Product_Types")
    ws_products = wb.create_sheet("Product_Types")
    
    products_header_fill = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
    
    ws_products.append(['PRODUCT TYPES'])
    ws_products.append(['Add different solar panel types if needed'])
    ws_products.append([''])
    ws_products.append(['Product Code', 'Product Name'])
    ws_products.append(['Monocrystalline', 'Monocrystalline Solar Panel'])
    ws_products.append(['', ''])
    
    # Format
    ws_products['A1'].font = Font(bold=True, size=14)
    ws_products['A2'].font = instruction_font
    for cell in ws_products[4]:
        cell.font = header_font
        cell.fill = products_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # ============================================================================
    # Adjust column widths for all sheets
    # ============================================================================
    print("  Formatting columns...")
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(filename)
    print(f"\n✓ Excel file created successfully: {filename}")
    print(f"✓ Total sheets: {len(wb.worksheets)}")
    print("\n📊 Sheets created:")
    for i, sheet in enumerate(wb.worksheets, 1):
        print(f"  {i:2d}. {sheet.title}")
    
    print("\n" + "="*70)
    print("SUCCESS! Now you can:")
    print("  1. Open supply_chain_data.xlsx")
    print("  2. Add/remove rows for multiple facilities")
    print("  3. Fill in your Germany data")
    print("  4. Run the optimization model")
    print("="*70)


if __name__ == '__main__':
    create_supply_chain_excel('supply_chain_data.xlsx')